import pandas as pd
import openpyxl
from datetime import datetime
from typing import List, Dict, Any

# Константы
LOG_FILE_IN = "result_output"
LOG_FILE_OUT = "./temp/clear_log.txt"
DATA_FILE_OUT = "./temp/dict_all.txt"

# Настройки вывода для DataFrame
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 300)

# Ключи для разных типов устройств
KEYS_MX = ['name', 'type', 'temp PEM_0', 'temp PEM_1', 'temp RE_0', 'temp RE_1',
           's_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5']
KEYS_ACX_4000 = ['name', 'type', 'temp PEM_0', 'temp PEM_1', 'temp RE_0', 's_fan_1', 's_fan_2']
KEYS_ACX_2100 = ['name', 'type', 'temp RE_0']


def clear_log() -> None:
    """Очищает входной лог-файл и создает очищенный файл."""
    with open(LOG_FILE_IN, "r") as f_in, open(LOG_FILE_OUT, "w") as f_out:
        for line in f_in:
            if "-----Outputs" in line:
                f_out.write(line)
            elif "Chassis" in line and len(line.split()) == 3 and "|match" not in line:
                f_out.write(line.split()[2] + "\n")
            elif "Temp  PEM 0" in line:
                f_out.write(line.split()[4] + "\n")
            elif "PEM " in line and "|match" not in line:
                f_out.write(line.split()[3] + "\n")
            elif "Routing Engine" in line and "CPU" not in line:
                split_line = line.split()
                if len(split_line) == 10:
                    f_out.write(split_line[3] + "\n")
                else:
                    f_out.write(split_line[4] + "\n")
            elif "Fan" in line:
                f_out.write(line.split()[3] + "\n")


def read_clear_data() -> List[Dict[str, Any]]:
    """Читает очищенный файл и преобразует данные в список словарей."""
    try:
        with open(LOG_FILE_OUT, "r") as file:
            lines = [line.strip() for line in file.readlines()]

        list_all = []
        current_device = None

        for line in lines:
            if "-----Outputs from " in line:
                current_device = line.split()[-2]
                list_all.append([current_device])
            elif current_device and line:
                list_all[-1].append(line)

        list_dict = []
        for item in list_all:
            if len(item) == 11:
                list_dict.append(dict(zip(KEYS_MX, item)))
            elif len(item) == 7:
                list_dict.append(dict(zip(KEYS_ACX_4000, item)))
            elif len(item) == 3:
                list_dict.append(dict(zip(KEYS_ACX_2100, item)))

        return list_dict

    except FileNotFoundError:
        print(f"Файл {LOG_FILE_OUT} не найден")
        return []
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return []


def save_data(data: List[Dict[str, Any]]) -> None:
    """Сохраняет данные в файл."""
    try:
        with open(DATA_FILE_OUT, "w") as f:
            for item in data:
                f.write(str(item) + "\n")
    except Exception as e:
        print(f"Ошибка при сохранении данных: {e}")


# Работа с датафреймом:
def analyze_data(df: pd.DataFrame) -> None:
    """Анализирует данные и выводит статистику."""
    # Подсчет устройств по типам
    cnt_chassi = df.groupby('type', as_index=False).size()

    # Подсчет устройств с отключенными вентиляторами
    fan_columns = ['s_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5']
    cnt_1_fan_check = (df[fan_columns] == 0).any(axis=1).sum()

    # Подсчет устройств с двумя отключенными вентиляторами (кроме MX104)
    cnt_2_fans_check = ((df['s_fan_1'] == 0) & (df['s_fan_2'] == 0) & (df["type"] != "MX104")).sum()

    # Подсчет устройств с высокой температурой
    temp_columns = ['temp PEM_0', 'temp PEM_1', 'temp RE_0', 'temp RE_1']
    temp_warm = ((df['temp PEM_0'] > 50) | (df['temp PEM_1'] > 50) | (df['temp RE_0'] > 50) | (
                df['temp RE_1'] > 50)).sum()

    # Фильтрация устройств с проблемами
    df_fan_alarm = df[(df['s_fan_1'] == 0) |
                      (df['s_fan_2'] == 0) |
                      (df['s_fan_3'] == 0) |
                      (df['s_fan_4'] == 0) |
                      (df['s_fan_5'] == 0)]

    df_temp_alarm = df[
        (df['temp PEM_0'] > 50) | (df['temp PEM_1'] > 50) | (df['temp RE_0'] > 50) | (df['temp RE_1'] > 50)]

    return {
        'device_count': cnt_chassi,
        'fans_disabled': cnt_1_fan_check,
        'two_fans_disabled': cnt_2_fans_check,
        'high_temp_devices': temp_warm,
        'fan_alarm_devices': df_fan_alarm,
        'temp_alarm_devices': df_temp_alarm}

# Создание отчетного файла
def save_to_excel_sheets(data_dict: dict) -> None:
    """Сохраняет данные в Excel файл с отдельными листами."""
    # Получить текущую дату
    current_date = str(datetime.now().strftime("%Y-%m-%d"))
    try:
        with pd.ExcelWriter("output "+current_date+".xlsx", engine='openpyxl') as writer:
            # Сохраняем DataFrame
            df.to_excel(writer,sheet_name="Общая информация", index=False)

            # Сохраняем статистику
            pd.DataFrame([{
                'Устройства с одним отключенным вентилятором': data_dict['fans_disabled'],
                'Устройства с двумя отключенными вентиляторами': data_dict['two_fans_disabled'],
                'Устройства с температурой > 50': data_dict['high_temp_devices']
            }]).to_excel(writer, sheet_name='Статистика', index=False)

            # Сохраняем устройства с отключенными вентиляторами
            data_dict['fan_alarm_devices'].to_excel(writer, sheet_name='Вентиляторы', index=False)

            # Сохраняем устройства с высокой температурой
            data_dict['temp_alarm_devices'].to_excel(writer, sheet_name='Температура', index=False)

            # Настройка стиля для всех листов
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # Установка ширины столбцов
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Установка выравнивания по центру для всех ячеек
                for column in worksheet.columns:
                    for cell in column:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            # Применение условного форматирования к основному листу
            worksheet = writer.sheets["Общая информация"]

            # Закрашивание ячеек с 0 в столбцах вентиляторов
            fan_columns = ['s_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5']
            for col in fan_columns:
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1  # +1 because openpyxl uses 1-based indexing
                    for row in range(2, len(df) + 2):  # +2 because of header and 1-based indexing
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value == 0:
                            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000",
                                                                    fill_type="solid")

            # Закрашивание ячеек с температурой > 50
            temp_columns = ['temp PEM_0', 'temp PEM_1', 'temp RE_0', 'temp RE_1']
            for col in temp_columns:
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        try:
                            if cell.value is not None and float(cell.value) > 50:
                                cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000",
                                                                        fill_type="solid")
                        except (ValueError, TypeError):
                            pass


        print("Данные успешно сохранены в output.xlsx")
    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {e}")


# Основная программа
if __name__ == "__main__":
    clear_log()
    data = read_clear_data()
    save_data(data)

    # Создание DataFrame для удобного просмотра
    if data:
        df = pd.DataFrame(data)
        print("Данные:")
        print(df)
        print()

        # Преобразование типов данных
        numeric_cols = ['temp PEM_0', 'temp PEM_1', 'temp RE_0', 'temp RE_1',
                        's_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5']

        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')

        # Анализ данных
        analysis_results = analyze_data(df)

        # Вывод статистики в консоль
        print("Количество устройств по типам:")
        print(analysis_results['device_count'])
        print()

        print(f"Устройства с отключенными вентиляторами: {analysis_results['fans_disabled']}")
        print(f"Устройства с двумя отключенными вентиляторами (не MX104): {analysis_results['two_fans_disabled']}")
        print(f"Устройства с температурой > 50: {analysis_results['high_temp_devices']}")
        print()

        print("Устройства с отключенными вентиляторами:")
        print(analysis_results['fan_alarm_devices'])
        print()

        print("Устройства с высокой температурой:")
        print(analysis_results['temp_alarm_devices'])
        print()

        # Сохранение в Excel
        save_to_excel_sheets(analysis_results)